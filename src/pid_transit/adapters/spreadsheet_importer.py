"""
Spreadsheet Importer Adapter.

Parses Excel (.xlsx) files or a directory of .csv files that follow the 
Transmodel relational schema and imports them into the TransmodelDatabase.
"""

import csv
import logging
from pathlib import Path
from typing import Dict, Any, Union

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ..core.database import TransmodelDatabase
from ..core.schemas import TRANSMODEL_ENTITIES
from ..core.exceptions import ImportFailedError

logger = logging.getLogger(__name__)

class SpreadsheetImporter:
    """
    Adapter class for importing data from raw CSV folders or Excel workbooks.
    Expects table structures to map 1:1 with the underlying Transmodel database.
    """
    
    def __init__(self, format: str = "csv"):
        """
        Args:
            format: 'csv' (expects a directory path) or 'xlsx' (expects a file path).
        """
        self.format = format.lower()
        if self.format not in ("csv", "xlsx"):
            raise ValueError("format must be 'csv' or 'xlsx'")

    def import_to_db(self, db: TransmodelDatabase, source_path: Union[Path, str]) -> Dict[str, int]:
        """
        Import the spreadsheet data into the database.
        
        Args:
            db: TransmodelDatabase instance.
            source_path: Path to the .xlsx file or the directory containing .csv files.
        """
        logger.info("Importing spreadsheet from %s (format=%s)", source_path, self.format)
        path = Path(source_path)
        stats = {}

        if self.format == "xlsx":
            if not HAS_OPENPYXL:
                raise ImportFailedError("openpyxl is required to import Excel files. Install via 'pip install openpyxl'")
            if not path.is_file() or path.suffix.lower() != ".xlsx":
                raise ImportFailedError(f"Expected a valid .xlsx file, got {path}")
            
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                for sheet_name in wb.sheetnames:
                    entity_name = sheet_name.lower().strip()
                    if entity_name not in TRANSMODEL_ENTITIES:
                        logger.warning("Skipping sheet '%s', not a recognized Transmodel entity.", sheet_name)
                        continue
                        
                    sheet = wb[sheet_name]
                    rows = list(sheet.iter_rows(values_only=True))
                    if len(rows) < 2:
                        continue
                        
                    headers = [str(h).strip() for h in rows[0] if h is not None]
                    records = []
                    for row in rows[1:]:
                        if not any(row):  # skip empty rows
                            continue
                        record = {}
                        for i, val in enumerate(row):
                            if i < len(headers):
                                # Convert all values to strings/None as per DB expectations,
                                # though Pydantic handles coercion.
                                record[headers[i]] = str(val).strip() if val is not None else None
                        records.append(record)
                        
                    if records:
                        inserted = db.upsert(entity_name, records)
                        stats[entity_name] = inserted
            except Exception as e:
                raise ImportFailedError(f"Failed to read Excel workbook: {e}") from e

        elif self.format == "csv":
            if not path.is_dir():
                raise ImportFailedError(f"For 'csv' format, expected a directory, got {path}")
            
            for file_path in path.glob("*.csv"):
                entity_name = file_path.stem.lower().strip()
                if entity_name not in TRANSMODEL_ENTITIES:
                    logger.warning("Skipping file '%s', not a recognized Transmodel entity.", file_path.name)
                    continue
                    
                try:
                    with open(file_path, "r", encoding="utf-8-sig") as f:
                        reader = csv.DictReader(f)
                        records = []
                        for row in reader:
                            # Clean up keys and values
                            clean_row = {}
                            for k, v in row.items():
                                if k:
                                    clean_row[k.strip()] = v.strip() if v else None
                            records.append(clean_row)
                            
                        if records:
                            inserted = db.upsert(entity_name, records)
                            stats[entity_name] = inserted
                except Exception as e:
                    raise ImportFailedError(f"Failed to read CSV {file_path.name}: {e}") from e
                    
        return stats
